from openpyxl import load_workbook

class ExcelUtil:
    def __init__(self, path, sheet_name):
        self.path = path
        self.workbook = load_workbook(path)
        self.sheet = self.workbook[sheet_name]

    def read_data(self, row, col):
        return self.sheet.cell(row=row, column=col).value

    def write_data(self, row, col, value):
        self.sheet.cell(row=row, column=col).value = value
        self.workbook.save(self.path)

    def get_row_count(self):
        return self.sheet.max_row
